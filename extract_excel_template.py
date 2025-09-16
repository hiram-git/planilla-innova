#!/usr/bin/env python3
"""
Script para extraer información de la plantilla Excel
"""
import sys
import os

# Intentar importar openpyxl
try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl no está instalado. Instalando...")
    os.system("pip install openpyxl")
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("No se pudo instalar openpyxl. Saliendo...")
        sys.exit(1)

def extract_excel_template(file_path):
    """Extraer información de la plantilla Excel"""
    try:
        # Cargar el workbook
        wb = load_workbook(file_path)
        
        print("=== INFORMACIÓN DE LA PLANTILLA EXCEL ===")
        print(f"Archivo: {file_path}")
        print(f"Hojas disponibles: {wb.sheetnames}")
        print()
        
        # Analizar cada hoja
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"--- HOJA: {sheet_name} ---")
            
            # Obtener dimensiones
            print(f"Dimensiones: {ws.max_row} filas x {ws.max_column} columnas")
            
            # Extraer headers y estructura (primeras 10 filas)
            print("\nPrimeras 10 filas de datos:")
            for row in range(1, min(11, ws.max_row + 1)):
                row_data = []
                for col in range(1, min(15, ws.max_column + 1)):  # Máximo 15 columnas
                    cell = ws.cell(row=row, column=col)
                    value = cell.value
                    if value is not None:
                        row_data.append(f"[{get_column_letter(col)}{row}]: {str(value)[:50]}")
                
                if row_data:
                    print(f"Fila {row}: {' | '.join(row_data)}")
            
            # Buscar celdas merged
            print(f"\nCeldas combinadas: {len(ws.merged_cells.ranges)}")
            for merged_range in ws.merged_cells.ranges:
                print(f"  - {merged_range}")
            
            # Analizar estilos de las primeras filas
            print("\nEstilos encontrados en las primeras filas:")
            for row in range(1, min(6, ws.max_row + 1)):
                for col in range(1, min(8, ws.max_column + 1)):
                    cell = ws.cell(row=row, column=col)
                    if cell.value and cell.fill.start_color.index != '00000000':
                        print(f"  {get_column_letter(col)}{row}: Color={cell.fill.start_color.index}, Font={cell.font.name}")
            
            print("\n" + "="*60)
        
    except Exception as e:
        print(f"Error al leer el archivo Excel: {e}")
        return False
    
    return True

if __name__ == "__main__":
    file_path = r"C:\xampp82\htdocs\planilla-claude-v2\assets\template\PRIMERA QUINCENA DE NOVIEMBRE.xlsx"
    
    if not os.path.exists(file_path):
        print(f"Error: El archivo {file_path} no existe")
        sys.exit(1)
    
    extract_excel_template(file_path)